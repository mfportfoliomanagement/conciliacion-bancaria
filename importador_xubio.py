# -*- coding: utf-8 -*-
"""
CAPA 3 — Generador del IMPORTADOR DE ASIENTOS para Xubio.

Toma los movimientos ya clasificados (Capa 2) y arma el asiento contable AGRUPADO
por CIRCUITO y por MES, con cada cuenta contra el banco, cerrando debe = haber.

Diseño confirmado leyendo +20 importadores reales del contador (VENEZUELA).
V1: agrupa por cuenta (sin discriminar por nombre en el auxiliar). Solo filas con importe > 0.

Genera dos hojas:
  - RESUMEN ASIENTOS  : hoja de trabajo (con control debe=haber por circuito).
  - IMPORTADOR ASIENTOS: la que se importa a Xubio (columnas fijas).

Regla de oro: las cuentas cuyo circuito no está definido con certeza NO se inventan;
van a un bloque "A REVISAR" para imputación humana (ej. AFIP-RENTAS, Caja, Moneda Extranjera).
"""
import calendar
import datetime
from collections import defaultdict, OrderedDict

# ---------------------------------------------------------------------------
# TABLA MAESTRA: cuenta interna (la que produce la Capa 2)  ->  (nombre EXACTO en Xubio, circuito)
# Los nombres de la derecha son los que vieron en los asientos reales del contador.
# ---------------------------------------------------------------------------
GB  = 'GASTOS BANCARIOS'
IMP = 'PAGO DE IMPUESTOS'
SUE = 'PAGO DE SUELDOS'
PRO = 'PAGO A PROVEEDORES'
COB = 'COBRANZAS Y DEPOSITOS'
RET = 'RETIROS Y CAJA CHICA'
INV = 'INVERSIONES EN BANCO'
OTR = 'OTROS MOVIMIENTOS'
TRP = 'TRANSFERENCIA ENTRE CUENTAS PROPIAS'
REVISAR = '__REVISAR__'   # sin circuito definido: queda para imputación humana

MAPA_CUENTA = {
    # cuenta interna (Capa 2)                 : (nombre exacto Xubio,            circuito)
    'Gastos bancarios':                         ('Gastos bancarios',                GB),
    'IVA Crédito Fiscal':                       ('IVA Crédito Fiscal',              GB),
    'Impuesto al Crédito Ley 25.413':           ('Impuesto al Crédito Ley 25.413',  GB),
    'Impuesto al Débito Ley 25.413':            ('Impuesto al Débito Ley 25.413',   GB),
    'Percepción Ingresos Brutos Sufrida':       ('Percepción Ingresos Brutos Sufrida', GB),
    'Percepción de IVA Sufrida':                ('Percepción de IVA Sufrida',       GB),
    'Sircreb':                                  ('Sircreb',                         GB),
    'Retenciones Ingresos Brutos CABA':         ('Retención Ingresos Brutos Sufrida', GB),

    'Impuestos y Tasas':                        ('Impuestos y Tasas',               IMP),

    'SUELDOS A PAGAR':                          ('SUELDOS A PAGAR',                 SUE),
    'Sueldos a pagar':                          ('SUELDOS A PAGAR',                 SUE),

    'Proveedores':                              ('Proveedores',                     PRO),
    'Acreedores Varios':                        ('Acreedores Varios',               PRO),

    'Deudores por Venta':                       ('Deudores por venta',              COB),
    'Deudores por venta':                       ('Deudores por venta',              COB),

    'EATON DIEGO MARTIN - Cuenta Particular':   ('EATON DIEGO MARTIN - Cuenta Particular', RET),
    'SANCHEZ EDUARDO OMAR - Cuenta Particular': ('SANCHEZ EDUARDO OMAR - Cuenta Particular', RET),

    'FCI':                                      ('FCI',                             INV),
    'Resultado por Inversión':                  ('Resultado por Inversión',         INV),

    'Agua':                                     ('AGUA',                            OTR),
    'Energía Eléctrica':                        ('ENERGÍA ELÉCTRICA',               OTR),
    'Telefonia':                                ('TELEFONÍA e INTERNET',            OTR),
    'ABL':                                      ('ABL',                             OTR),
    'Expensas':                                 ('Expensas',                        OTR),   # nombre a confirmar en Xubio
    'Seguros':                                  ('Seguros',                         OTR),   # nombre a confirmar en Xubio
    'GASTOS VARIOS':                            ('GASTOS VARIOS',                   OTR),
    'SEGURIDAD':                                ('SEGURIDAD',                       OTR),

    'Transferencias entre cuentas':             ('Transferencias entre cuentas',    TRP),

    # --- Sin circuito definido: NO se imputan solas, van a revisión humana ---
    'AFIP-RENTAS':                              ('AFIP-RENTAS',                     REVISAR),  # se desglosa con el VEP
    'Caja':                                     ('Caja',                            REVISAR),  # ajuste manual del contador
    'Moneda Extranjera':                        ('Moneda Extranjera',               REVISAR),
}

# Orden en que se listan los circuitos en la salida (como en los importadores reales)
ORDEN_CIRCUITOS = [GB, RET, TRP, IMP, SUE, PRO, COB, INV, OTR]

# Nombre de la cuenta contable del BANCO por número de cuenta (parametrizable).
# La clave es un fragmento que aparezca en res.cuenta. Si no matchea ninguno -> BANCO_DEFAULT.
BANCO_POR_CUENTA = {
    '8995': 'Banco Galicia en $',   # cuenta Galicia 0008995
    '1941': 'Banco',                # cuenta Galicia 0001941
}
BANCO_DEFAULT = 'Banco'


def _nombre_banco(res) -> str:
    cta = str(getattr(res, 'cuenta', '') or '')
    for frag, nombre in BANCO_POR_CUENTA.items():
        if frag in cta:
            return nombre
    return BANCO_DEFAULT


def _ultimo_dia_mes(anio: int, mes: int) -> datetime.date:
    return datetime.date(anio, mes, calendar.monthrange(anio, mes)[1])


def _r2(x) -> float:
    return round(float(x or 0.0), 2)


def construir_asientos(resultados):
    """Devuelve una lista de asientos (uno por circuito y mes y cuenta bancaria).
    Cada asiento: dict(fecha, concepto, banco, circuito, lineas=[(cuenta, debe, haber)], revisar=bool).
    """
    asientos = []
    revisar_global = []  # movimientos sin cuenta o con circuito REVISAR

    for res in resultados:
        banco = _nombre_banco(res)
        etiqueta_cta = str(getattr(res, 'cuenta', '') or '').strip()

        # agrupador[(anio, mes, circuito)][nombre_xubio] = [debe, haber]
        agrupador = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))

        for m in res.movimientos:
            cuenta_int = getattr(m, 'cuenta_sugerida', '') or ''
            fecha = m.fecha
            if fecha is None:
                revisar_global.append((res, m, 'sin fecha'))
                continue
            clave_mes = (fecha.year, fecha.month)

            if not cuenta_int:
                revisar_global.append((res, m, 'sin cuenta'))
                continue
            mapeo = MAPA_CUENTA.get(cuenta_int)
            if mapeo is None:
                revisar_global.append((res, m, f'cuenta no mapeada: {cuenta_int}'))
                continue
            nombre_xubio, circuito = mapeo
            if circuito == REVISAR:
                revisar_global.append((res, m, f'circuito a revisar: {cuenta_int}'))
                continue

            slot = agrupador[(clave_mes[0], clave_mes[1], circuito)][nombre_xubio]
            slot[0] = _r2(slot[0] + _r2(m.debito))
            slot[1] = _r2(slot[1] + _r2(m.credito))

        # armar los asientos ordenados por mes y por ORDEN_CIRCUITOS
        claves = sorted(agrupador.keys(), key=lambda k: (k[0], k[1], ORDEN_CIRCUITOS.index(k[2]) if k[2] in ORDEN_CIRCUITOS else 99))
        for (anio, mes, circuito) in claves:
            cuentas = agrupador[(anio, mes, circuito)]
            fecha = _ultimo_dia_mes(anio, mes)
            suf = f"{mes:02d}.{anio}"
            concepto = f"{circuito} - BCO. {banco} {etiqueta_cta} del mes {suf}".replace('  ', ' ')

            # NETEADO: una sola línea por cuenta (debe - haber). Si el neto queda en 0,
            # la cuenta se cancela sola y no se lista.
            lineas = []
            neto_cuentas = 0.0   # suma de (debe - haber) de todas las cuentas del circuito
            for nombre_xubio, (debe, haber) in cuentas.items():
                neto = _r2(debe - haber)
                if neto > 0:
                    lineas.append((nombre_xubio, neto, 0.0))
                elif neto < 0:
                    lineas.append((nombre_xubio, 0.0, _r2(-neto)))
                neto_cuentas = _r2(neto_cuentas + neto)

            # Contrapartida banco (también neteada), para que el asiento cierre
            if neto_cuentas > 0:
                lineas.append((banco, 0.0, neto_cuentas))
            elif neto_cuentas < 0:
                lineas.append((banco, _r2(-neto_cuentas), 0.0))

            if lineas:
                asientos.append(dict(fecha=fecha, concepto=concepto, banco=banco,
                                     circuito=circuito, lineas=lineas))
    return asientos, revisar_global


def _control_asiento(asiento):
    d = _r2(sum(l[1] for l in asiento['lineas']))
    h = _r2(sum(l[2] for l in asiento['lineas']))
    return d, h, _r2(d - h)


def control_cobertura(resultados):
    """CONTROL GLOBAL de la Capa 3.
    Compara, por cada cuenta bancaria, cuánto movió el banco en el EXTRACTO
    (créditos - débitos de TODOS los movimientos) contra cuánto quedó reflejado
    en el IMPORTADOR (solo los movimientos que SÍ se pudieron clasificar).
    Si la diferencia no es 0, el importador está INCOMPLETO por ese monto:
    hay movimientos sin cuenta que faltan imputar (ver hoja A REVISAR).
    Devuelve una lista de dicts, uno por cuenta.
    """
    reporte = []
    for res in resultados:
        total_extracto = 0.0      # todo lo que movió el banco
        total_importado = 0.0     # lo que entró al importador (clasificado y con circuito)
        falta_monto = 0.0
        falta_cant = 0
        for m in res.movimientos:
            neto = _r2(_r2(m.credito) - _r2(m.debito))
            total_extracto = _r2(total_extracto + neto)
            cuenta_int = getattr(m, 'cuenta_sugerida', '') or ''
            mapeo = MAPA_CUENTA.get(cuenta_int)
            clasificado_ok = bool(cuenta_int) and mapeo is not None and mapeo[1] != REVISAR and m.fecha is not None
            if clasificado_ok:
                total_importado = _r2(total_importado + neto)
            else:
                falta_monto = _r2(falta_monto + abs(neto))
                falta_cant += 1
        reporte.append(dict(
            cuenta=str(getattr(res, 'cuenta', '') or ''),
            total_extracto=total_extracto,
            total_importado=total_importado,
            diferencia=_r2(total_extracto - total_importado),
            faltan_movimientos=falta_cant,
            faltan_monto=falta_monto,
            completo=(falta_cant == 0),
        ))
    return reporte


def generar_excel(resultados, path):
    """Genera el .xlsx con las hojas RESUMEN ASIENTOS e IMPORTADOR ASIENTOS."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    asientos, revisar = construir_asientos(resultados)
    wb = openpyxl.Workbook()

    # ---------- Hoja IMPORTADOR ASIENTOS ----------
    ws = wb.active
    ws.title = 'IMPORTADOR ASIENTOS'
    cols = ['FECHA', 'CONCEPTO', 'CIRCUITOCONTABLE', 'CUENTA', 'DEBE', 'HABER',
            'ORGANIZACION', 'CENTRODECOSTO', 'DESCRIPCION']
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(1, c).font = Font(bold=True)
    for a in asientos:
        ws.append([a['fecha'].strftime('%d/%m/%Y'), a['concepto'], 'default', '', '', '', '', '', ''])
        for (cuenta, debe, haber) in a['lineas']:
            ws.append(['', '', '', cuenta,
                       debe if debe > 0 else '', haber if haber > 0 else '', '', '', ''])
    for col, w in zip('ABCDEFGHI', [12, 55, 16, 40, 15, 15, 14, 14, 30]):
        ws.column_dimensions[col].width = w

    # ---------- Hoja RESUMEN ASIENTOS ----------
    wr = wb.create_sheet('RESUMEN ASIENTOS')
    wr.append(['FECHA', 'CONCEPTO', 'NOMBRE CUENTA', 'DEBE', 'HABER', 'CONTROL'])
    for c in range(1, 7):
        wr.cell(1, c).font = Font(bold=True)
    amarillo = PatternFill('solid', fgColor='FFF2CC')
    for a in asientos:
        d, h, ctrl = _control_asiento(a)
        wr.append([a['fecha'].strftime('%d/%m/%Y'), a['concepto'], '', '', '', ''])
        for (cuenta, debe, haber) in a['lineas']:
            wr.append(['', '', cuenta, debe if debe > 0 else '', haber if haber > 0 else '', ''])
        fila_total = ['', '', 'TOTAL', d, h, ctrl]
        wr.append(fila_total)
        r = wr.max_row
        for c in range(3, 7):
            wr.cell(r, c).font = Font(bold=True)
            wr.cell(r, c).fill = amarillo
    for col, w in zip('ABCDEF', [12, 55, 40, 15, 15, 12]):
        wr.column_dimensions[col].width = w

    # ---------- Hoja A REVISAR (movimientos sin circuito) ----------
    if revisar:
        wv = wb.create_sheet('A REVISAR')
        wv.append(['CUENTA (extracto)', 'FECHA', 'CONCEPTO', 'NOMBRE', 'DÉBITO', 'CRÉDITO', 'MOTIVO'])
        for c in range(1, 8):
            wv.cell(1, c).font = Font(bold=True)
        for (res, m, motivo) in revisar:
            wv.append([str(getattr(res, 'cuenta', '')),
                       m.fecha.strftime('%d/%m/%Y') if m.fecha else '',
                       m.concepto, getattr(m, 'nombre', ''),
                       _r2(m.debito) or '', _r2(m.credito) or '', motivo])
        for col, w in zip('ABCDEFG', [18, 12, 40, 28, 14, 14, 30]):
            wv.column_dimensions[col].width = w

    # ---------- Hoja CONTROL (cobertura global) ----------
    cob = control_cobertura(resultados)
    wc = wb.create_sheet('CONTROL')
    wc.append(['CUENTA', 'MOVIÓ EL BANCO (extracto)', 'REFLEJADO EN IMPORTADOR',
               'DIFERENCIA (falta imputar)', 'MOV. SIN CLASIFICAR', 'ESTADO'])
    for c in range(1, 7):
        wc.cell(1, c).font = Font(bold=True)
    verde = PatternFill('solid', fgColor='C6EFCE')
    rojo = PatternFill('solid', fgColor='FFC7CE')
    for r in cob:
        estado = 'COMPLETO ✓' if r['completo'] else 'INCOMPLETO — faltan cuentas'
        wc.append([r['cuenta'], r['total_extracto'], r['total_importado'],
                   r['diferencia'], r['faltan_movimientos'], estado])
        fila = wc.max_row
        wc.cell(fila, 6).fill = verde if r['completo'] else rojo
        wc.cell(fila, 6).font = Font(bold=True)
    for col, w in zip('ABCDEF', [22, 26, 26, 26, 20, 30]):
        wc.column_dimensions[col].width = w

    wb.save(path)
    return asientos, revisar
