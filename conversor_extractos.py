# -*- coding: utf-8 -*-
"""
============================================================================
 CONVERSOR DE EXTRACTOS BANCARIOS  ->  TABLA ESTÁNDAR PARA CONCILIACIÓN
============================================================================

Qué hace
--------
Toma el PDF de un extracto bancario (descargado del homebanking, con texto
real adentro) y lo convierte en una tabla estándar de movimientos, una fila
por movimiento, con estas columnas:

    fecha | concepto | referencia | cuit | nombre | debito | credito | saldo

Esa tabla es la que después llena la hoja "Carga de extracto" del archivo de
conciliación. Cada banco arma su PDF distinto, así que hay UN LECTOR POR BANCO,
pero todos devuelven el mismo objeto Resultado con la misma tabla estándar.

Diseño (para que sea mantenible y sobreviva a cambios de formato)
-----------------------------------------------------------------
  * Cada lector es una función leer_<banco>(pdf) -> Resultado.
  * detectar_banco() mira el texto del PDF y elige el lector.
  * Toda la lógica frágil (posiciones X de columnas, rarezas de cada banco)
    está aislada y COMENTADA dentro de su lector. Si un banco cambia el PDF,
    se toca solo ese lector.
  * RED DE SEGURIDAD: verificar_control() reconstruye el saldo por acumulación
    (saldo_ini + créditos - débitos) y lo compara contra el saldo final que
    informa el banco. Si da 0, no se perdió ni se duplicó ningún movimiento.

Uso
---
    python conversor_extractos.py  archivo.pdf  [salida.xlsx]

    # o desde código:
    from conversor_extractos import leer_extracto, exportar_excel
    res = leer_extracto("extracto.pdf")
    exportar_excel(res, "convertido.xlsx")

Bancos soportados en este módulo: Banco Ciudad, Banco Comafi, Supervielle.
Pendientes de sumar (dejar su función leer_<banco> siguiendo el mismo patrón):
Galicia, BBVA, Macro, ICBC, Credicoop.
============================================================================
"""
from dataclasses import dataclass, field
from typing import Optional, List
import re, datetime, sys
import pdfplumber

# ---------------------------------------------------------------------------
# Modelo de datos estándar (la "interfaz" común que consume la conciliación)
# ---------------------------------------------------------------------------
@dataclass
class Movimiento:
    fecha: datetime.date
    concepto: str
    referencia: str = ''
    cuit: str = ''
    nombre: str = ''
    debito: float = 0.0
    credito: float = 0.0
    saldo: Optional[float] = None          # None si el banco no lo imprime en esa fila
    categoria: str = ''                    # p.ej. 'SUELDOS' (clasificación); '' = sin clasificar
    cuenta_sugerida: str = ''              # cuenta contable sugerida para el importador Xubio

@dataclass
class Resultado:
    banco: str
    titular: str
    cuit_titular: str
    cuenta: str
    periodo: str
    saldo_ini: float
    saldo_fin: float
    movimientos: List[Movimiento] = field(default_factory=list)

# ---------------------------------------------------------------------------
# Helpers compartidos
# ---------------------------------------------------------------------------
def num_ar(txt: str) -> float:
    """Convierte número en formato argentino a float.
       '1.234.567,89' -> 1234567.89 ; signo negativo puede ir al final ('...,72-')."""
    txt = txt.strip()
    neg = txt.endswith('-')
    txt = txt.rstrip('-')
    return (-1 if neg else 1) * float(txt.replace('.', '').replace(',', '.'))

RE_NUM = re.compile(r'^-?[\d.]+,\d{2}-?$')

def fecha_dmy2(txt: str):
    """dd/mm/yy -> date."""
    m = re.match(r'(\d{2})/(\d{2})/(\d{2})$', txt)
    return datetime.date(2000+int(m[3]), int(m[2]), int(m[1])) if m else None

def fecha_dmy4(txt: str):
    """dd/mm/yyyy -> date (año completo; lo usa el extracto Office Banking de Galicia)."""
    m = re.match(r'(\d{2})/(\d{2})/(\d{4})$', txt)
    return datetime.date(int(m[3]), int(m[2]), int(m[1])) if m else None

_MESES = {'ENE':1,'FEB':2,'MAR':3,'ABR':4,'MAY':5,'JUN':6,
          'JUL':7,'AGO':8,'SEP':9,'OCT':10,'NOV':11,'DIC':12}
def fecha_dMMMy(txt: str):
    """dd-MMM-yyyy (mes en letras, español) -> date."""
    m = re.match(r'(\d{2})-([A-Z]{3})-(\d{4})', txt.upper())
    return datetime.date(int(m[3]), _MESES[m[2]], int(m[1])) if m and m[2] in _MESES else None

def _filas_por_top(words, y_min, y_max, tol=2):
    """Agrupa palabras en filas por su coordenada 'top', con tolerancia vertical."""
    filas = {}
    for w in words:
        if not (y_min <= w['top'] <= y_max):
            continue
        key = round(w['top'])
        for k in list(filas):
            if abs(k-key) <= tol:
                key = k; break
        filas.setdefault(key, []).append(w)
    return [sorted(filas[k], key=lambda z: z['x0']) for k in sorted(filas)]

# ===========================================================================
# LECTOR: BANCO CIUDAD
# ---------------------------------------------------------------------------
# Formato prolijo, una línea por movimiento, saldo acumulado en CADA fila.
# CUIT/nombre vienen EN la misma línea (columna "Descripción de movimiento").
# Trampa propia: marca de agua vertical en el margen izquierdo (x0<28) que se
# cuela en la columna de fecha -> se descarta por posición.
# Columnas (x): fecha<82 | concepto<170 | desc>=370 | números por borde der.:
#               débito x1<236 | crédito x1<300 | saldo resto.
# ===========================================================================
def leer_ciudad(pdf) -> Resultado:
    p = pdf.pages[0]
    W = p.extract_words()
    def num_box(y0, y1, x0min):
        for w in W:
            if y0 < w['top'] < y1 and w['x0'] > x0min and RE_NUM.match(w['text']):
                return num_ar(w['text'])
        return None
    saldo_ini = num_box(195, 202, 460)
    saldo_fin = num_box(780, 786, 460)

    def col(x0, x1):
        if x0 < 82:  return 'fecha'
        if x0 < 170: return 'concepto'
        if x0 >= 370:return 'desc'
        if x1 < 236: return 'debito'
        if x1 < 300: return 'credito'
        return 'saldo'

    movs = []
    for fila in _filas_por_top([w for w in W if w['x0'] >= 28], 225, 400):
        celdas = {k: [] for k in ('fecha','concepto','debito','credito','saldo','desc')}
        for w in fila:
            celdas[col(w['x0'], w['x1'])].append(w['text'])
        fecha = fecha_dMMMy(' '.join(celdas['fecha']))
        if not fecha:
            continue
        desc = ' '.join(celdas['desc']).strip()
        mc = re.match(r'(\d{11})[-\s]*(.*)', desc)
        cuit, nombre = (mc[1], mc[2].strip()) if mc else ('', desc)
        movs.append(Movimiento(
            fecha=fecha, concepto=' '.join(celdas['concepto']).strip(),
            cuit=cuit, nombre=nombre,
            debito=num_ar(celdas['debito'][0]) if celdas['debito'] else 0.0,
            credito=num_ar(celdas['credito'][0]) if celdas['credito'] else 0.0,
            saldo=num_ar(celdas['saldo'][0]) if celdas['saldo'] else None))

    txt = p.extract_text()
    tit = re.search(r'\n(ASC .+?)\n', txt)
    cuit = re.search(r'(\d{2}-\d{8}-\d)', txt)
    cta  = re.search(r'(\d{7}/\d)', txt)
    return Resultado('Banco Ciudad', tit[1].strip() if tit else '',
                     cuit[1] if cuit else '', cta[1] if cta else '',
                     '', saldo_ini, saldo_fin, movs)

# ===========================================================================
# LECTOR: BANCO COMAFI
# ---------------------------------------------------------------------------
# Multi-página con líneas "Transporte" que arrastran el saldo entre hojas
# (se usan como límite de banda, NO son movimientos). El saldo NO aparece en
# cada fila. Movimientos que se parten en 2 líneas (ej "Pago de cheque": ref y
# nombre en un renglón, importe en el siguiente). Saldo negativo con '-' final.
# CUIT/nombre en sección aparte "Transferencias Electrónicas" -> cruce por
# fecha+importe (estilo BBVA), excluyendo los internos del banco (30604731018).
# Trampa: la palabra "Fecha" reaparece abajo (sección CHEQUES) -> el encabezado
# solo se busca en la zona de la tabla (top<270).
# Columnas (x): fecha<64 | concepto<196 | ref<290 | números por borde der.:
#               débito x1<430 | crédito x1<510 | saldo resto.
# ===========================================================================
def _comafi_transferencias(pdf):
    txt = pdf.pages[5].extract_text().split('\n')
    envs, modo = [], None
    for ln in txt:
        if 'ENVIADAS' in ln:  modo = 'env'; continue
        if 'RECIBIDAS' in ln: modo = 'rec'; continue
        if 'FONDOS COMUNES' in ln: break
        m = re.match(r'(\d{2}/\d{2}/\d{4})\s+(.*)', ln.strip())
        if not m or modo is None: continue
        imp  = re.search(r'\$\s*([\d.]+,\d{2})', m[2])
        cuit = re.search(r'\b(\d{11})\b', m[2])
        if not imp: continue
        cu = cuit[1] if cuit else ''
        if cu == '30604731018':       # movimientos internos (comisiones del banco)
            continue
        nombre = m[2][cuit.end():imp.start()].strip() if cuit else ''
        envs.append((datetime.datetime.strptime(m[1], '%d/%m/%Y').date(),
                     num_ar(imp[1]), cu, nombre))
    return envs

def leer_comafi(pdf) -> Resultado:
    envs = _comafi_transferencias(pdf)
    def col_num(x1): return 'debito' if x1 < 430 else 'credito' if x1 < 510 else 'saldo'
    saldo_ini = saldo_fin = None
    movs: List[Movimiento] = []

    for pg in range(1, 5):
        p = pdf.pages[pg]
        W = [w for w in p.extract_words() if not set(w['text']) <= set('-')]
        tops_hdr = [w['top'] for w in W if w['text'] == 'Fecha' and w['top'] < 270]
        tops_tr  = [w['top'] for w in W if w['text'] == 'Transporte']
        tops_al  = [w['top'] for w in W if w['text'] == 'al:']
        start = max(tops_hdr + [t for t in tops_tr if t < 270] or [150])
        fin_c = [t for t in tops_tr if t > 300] + tops_al
        end   = min(fin_c) if fin_c else 720

        for fila in _filas_por_top(W, start+2, end-1):
            textos = [w['text'] for w in fila]
            if 'Conceptos' in textos or 'Referencias' in textos:
                continue
            if any(w['text'] == 'Anterior' for w in fila):        # Saldo Anterior
                nums = [w for w in fila if RE_NUM.match(w['text'])]
                if nums: saldo_ini = num_ar(nums[-1]['text'])
                continue
            fecha_tok = [w for w in fila if w['x0'] < 64 and fecha_dmy2(w['text'])]
            concepto = ' '.join(w['text'] for w in fila if 64 <= w['x0'] < 196)
            ref = ' '.join(w['text'] for w in fila if 196 <= w['x0'] < 290 and w['x1'] < 290)
            nums = {'debito': None, 'credito': None, 'saldo': None}
            for w in fila:
                if RE_NUM.match(w['text']) and w['x0'] > 355:
                    nums[col_num(w['x1'])] = num_ar(w['text'])
            if fecha_tok:
                movs.append(Movimiento(
                    fecha=fecha_dmy2(fecha_tok[0]['text']), concepto=concepto.strip(),
                    referencia=ref.strip(), debito=nums['debito'] or 0.0,
                    credito=nums['credito'] or 0.0, saldo=nums['saldo']))
            elif movs:                                            # continuación
                extra = ' '.join(w['text'] for w in fila if w['x0'] < 290)
                if extra: movs[-1].concepto = (movs[-1].concepto + ' ' + extra).strip()
                if nums['debito']:  movs[-1].debito  += nums['debito']
                if nums['credito']: movs[-1].credito += nums['credito']
                if nums['saldo'] is not None: movs[-1].saldo = nums['saldo']

    for w in pdf.pages[4].extract_words():
        if w['text'] == 'al:':
            fila = [x for x in pdf.pages[4].extract_words() if abs(x['top']-w['top']) < 3]
            nums = [x for x in fila if RE_NUM.match(x['text'])]
            if nums: saldo_fin = num_ar(nums[-1]['text'])

    for m in movs:                                               # enriquecer CUIT
        if 'Transferencia Terceros' in m.concepto or ('Transf' in m.concepto and 'sueldos' in m.concepto):
            cand = [e for e in envs if e[0] == m.fecha and abs(e[1]-m.debito) < 0.01]
            if len(cand) == 1:
                m.cuit, m.nombre = cand[0][2], cand[0][3]

    t = pdf.pages[1].extract_text()
    tit  = re.search(r'TITULAR.*?\n.*?\n\s*(.+?)\s+\d{2}-\d{8}-\d', t, re.S)
    cuit = re.search(r'(\d{2}-\d{8}-\d)\s+Responsable', t)
    cta  = re.search(r'(\d{4}-\d{5}-\d)', t)
    return Resultado('Banco Comafi', (tit[1].strip() if tit else 'NEXO IT SRL'),
                     cuit[1] if cuit else '', cta[1] if cta else '',
                     '', saldo_ini, saldo_fin, movs)

# ===========================================================================
# LECTOR: SUPERVIELLE
# ---------------------------------------------------------------------------
# Una página de movimientos, texto limpio. Movimientos con líneas de
# continuación (Operación.../Cuentas Propias/CUIT-nombre) que NO traen importe.
# Trampa: la línea "Imp Ley 25413 s/Debitos" trae un importe en débito pero
# NO tiene saldo y NO mueve el balance (es informativa) -> se descarta.
# Regla: fila sin fecha + con importe pero sin saldo = informativa (descartar);
#        fila sin fecha + sin importe = continuación de texto del mov. previo.
# Columnas (x): fecha<100 | concepto<250 | ref<306 | números (x0>305) por borde
#               der.: débito x1<420 | crédito x1<470 | saldo resto.
# ===========================================================================
def leer_supervielle(pdf) -> Resultado:
    p = pdf.pages[0]
    W = [w for w in p.extract_words() if not set(w['text']) <= set('*')]
    def col_num(x1): return 'debito' if x1 < 420 else 'credito' if x1 < 470 else 'saldo'
    top_ini = min([w['top'] for w in W if w['text'] == 'anterior'] or [270])
    top_fin = max([w['top'] for w in W if w['text'] == 'ACTUAL'] or [620])

    saldo_ini = saldo_fin = None
    movs: List[Movimiento] = []
    for fila in _filas_por_top(W, top_ini-2, top_fin+2):
        txt = ' '.join(w['text'] for w in fila)
        nums = {'debito': None, 'credito': None, 'saldo': None}
        for w in fila:
            if RE_NUM.match(w['text']) and w['x0'] > 305:
                nums[col_num(w['x1'])] = num_ar(w['text'])
        if 'período anterior' in txt: saldo_ini = nums['saldo']; continue
        if 'PERIODO ACTUAL' in txt:   saldo_fin = nums['saldo']; continue
        fecha_tok = [w for w in fila if w['x0'] < 100 and fecha_dmy2(w['text'])]
        concepto = ' '.join(w['text'] for w in fila if 100 <= w['x0'] < 250)
        ref = ' '.join(w['text'] for w in fila if 250 <= w['x0'] < 306 and w['x1'] < 306)
        if fecha_tok:
            movs.append(Movimiento(
                fecha=fecha_dmy2(fecha_tok[0]['text']), concepto=concepto.strip(),
                referencia=ref.strip(), debito=nums['debito'] or 0.0,
                credito=nums['credito'] or 0.0, saldo=nums['saldo'], nombre=''))
            movs[-1]._cont = ''
        else:
            if nums['debito'] or nums['credito']:       # informativa -> descartar
                continue
            if movs:                                    # continuación de texto
                extra = ' '.join(w['text'] for w in fila if w['x0'] >= 100)
                movs[-1]._cont = (getattr(movs[-1], '_cont', '') + ' ' + extra).strip()

    for m in movs:
        mc = re.search(r'\b(\d{11})\b\s*(.*)', getattr(m, '_cont', ''))
        if mc: m.cuit, m.nombre = mc[1], mc[2].strip()

    t = p.extract_text()
    tit  = re.search(r'([A-ZÑ ]*OXXON[A-ZÑ ]*)', t)
    cuit = re.search(r'C\.U\.I\.T\.\s*0?(\d{2}-\d{8}-\d)', t)
    cta  = re.search(r'Nro\.:\s*([\d-]+)', t)
    return Resultado('Supervielle', (tit[1].strip() if tit else ''),
                     cuit[1] if cuit else '', cta[1] if cta else '',
                     '', saldo_ini, saldo_fin, movs)

# ===========================================================================
# LECTORES REINTEGRADOS: GALICIA, BBVA, MACRO
# ---------------------------------------------------------------------------
# Portados desde el desarrollo original. Galicia = 1 cuenta por PDF.
# BBVA y Macro = VARIAS cuentas por PDF -> devuelven una LISTA de Resultado.
# Agrupación de líneas por cercanía vertical (tol=4) para no perder importes
# por micro-desalineación (fue el bug que en BBVA hacía perder movimientos).
# ===========================================================================
_CUIT11 = re.compile(r'^\d{11}$')

def _cluster(words, tol=4):
    ws = sorted(words, key=lambda w: w['top']); out=[]; cur=[]; ref=None
    for w in ws:
        if ref is None or abs(w['top']-ref) <= tol:
            cur.append(w)
            if ref is None: ref=w['top']
        else:
            out.append(cur); cur=[w]; ref=w['top']
    if cur: out.append(cur)
    return [sorted(l, key=lambda z: z['x0']) for l in out]

def _fecha_ddmm(txt, anio):
    m = re.match(r'(\d{2})/(\d{2})$', txt)
    return datetime.date(int(anio), int(m[2]), int(m[1])) if m else None

# ---------------------------------------------------------------------------
# GALICIA — bloques multilínea (nombre / CUIT / CBU apilados bajo el movimiento)
# Columnas por borde derecho: crédito x1<400 | débito x1<500 | saldo resto.
# Totales y saldo final en la fila "Total". Saldo negativo con "-" al final.
# ---------------------------------------------------------------------------
def leer_galicia(pdf) -> Resultado:
    DATE = re.compile(r'^\d{2}/\d{2}/\d{2}$')
    def zona(x1): return 'credito' if x1 < 400 else ('debito' if x1 < 500 else 'saldo')
    t0 = pdf.pages[0].extract_text() or ''
    mcu = re.search(r'Responsable Impositivo\s*:\s*([\d\-]+)', t0)
    cuit_tit = mcu.group(1) if mcu else ''
    mct = re.search(r'N[°º]\s*([\d\-]+\s*[\d\-]+)', t0)
    cuenta = mct.group(1).strip() if mct else ''
    fechas = re.findall(r'\d{2}/\d{2}/\d{4}', t0)
    periodo = f"{fechas[0]} a {fechas[1]}" if len(fechas) >= 2 else ''
    mtt = re.search(r'\n([A-ZÑ0-9&\.\- ]{3,40})\s*\nResumen de Cuenta', t0)
    titular = mtt.group(1).strip() if mtt else ''

    movs=[]; cur=None; ended=False; tot=None
    for page in pdf.pages:
        for toks in _cluster(page.extract_words()):
            txt = ' '.join(w['text'] for w in toks).strip()
            if txt.startswith('Total'):
                amts=[num_ar(w['text']) for w in toks if RE_NUM.match(w['text'])]
                if len(amts) >= 3: tot=(abs(amts[0]), abs(amts[1]), amts[2])
                if cur: movs.append(cur); cur=None
                ended=True; continue
            if ended: continue
            first=toks[0]
            if DATE.match(first['text']) and first['x0'] < 80:
                if cur: movs.append(cur)
                cur=dict(fecha=first['text'], desc=[], credito=0.0, debito=0.0, saldo=None, det=[])
                for w in toks[1:]:
                    x0,x1,tx = w['x0'], w['x1'], w['text']
                    if RE_NUM.match(tx):
                        v=num_ar(tx); z=zona(x1)
                        if z=='credito': cur['credito']=abs(v)
                        elif z=='debito': cur['debito']=abs(v)
                        else: cur['saldo']=v
                    elif 80 <= x0 < 220:
                        cur['desc'].append(tx)
            else:
                if cur is None: continue
                det=[w['text'] for w in toks if 80 <= w['x0'] < 300]
                if det: cur['det'].append(' '.join(det))
    if cur and not ended: movs.append(cur)

    out=[]
    for m in movs:
        det=m['det']; nombre=det[0] if det else ''
        cu=next((d.strip() for d in det if _CUIT11.match(d.strip())), '')
        out.append(Movimiento(fecha_dmy2(m['fecha']), ' '.join(m['desc']).strip(),
                              '', cu, nombre, m['debito'], m['credito'], m['saldo']))
    saldo_fin = tot[2] if tot else (out[-1].saldo if out else 0.0)
    saldo_ini = round(out[0].saldo - (out[0].credito - out[0].debito), 2) if out and out[0].saldo is not None else 0.0
    return Resultado('Banco Galicia', titular, cuit_tit, cuenta, periodo, saldo_ini, saldo_fin, out)

# ---------------------------------------------------------------------------
# GALICIA "OFFICE BANKING" (export "Movimientos de CC") — FORMATO DISTINTO al
# "Resumen de Cuenta Corriente" clásico. Diferencias que obligan a un lector propio:
#   * Columnas en orden Fecha | Descripción | DÉBITOS | CRÉDITOS | Saldos
#     (débito ANTES que crédito; en el clásico es al revés).
#   * Fecha con AÑO COMPLETO (dd/mm/aaaa), no dd/mm/aa.
#   * Importe con SIGNO: "+ $ ..." = crédito ; "- $ ..." = débito. El saldo va sin signo.
#   * NO hay fila "Total" final. El saldo aparece en CADA fila -> el control de saldo
#     fila por fila es la garantía de integridad.
#   * No trae nombre del titular ni su CUIT en la cabecera (quedan vacíos).
# Posiciones reales medidas sobre el PDF (borde derecho x1 de cada número):
#   débito  ~329 | crédito ~428 | saldo ~534 . Se clasifica por el SIGNO (robusto),
#   con respaldo por posición si faltara el signo.
# ---------------------------------------------------------------------------
def leer_galicia_ob(pdf) -> Resultado:
    from collections import defaultdict
    HDR = {'Fecha', 'Descripción', 'Débitos', 'Créditos', 'Saldos'}
    DATE4 = re.compile(r'^\d{2}/\d{2}/\d{4}$')

    # Cabecera de la cuenta (solo en página 1)
    t0 = pdf.pages[0].extract_text() or ''
    mcta = re.search(r'Movimientos de CC.*?\$?\s*([\d\-]+\s+[\d\-]+)', t0)
    cuenta = mcta.group(1).strip() if mcta else ''

    movs = []; cur = None
    for page in pdf.pages:
        lineas = defaultdict(list)
        for w in page.extract_words():
            lineas[round(w['top'])].append(w)
        for top in sorted(lineas):
            ws = sorted(lineas[top], key=lambda w: w['x0'])
            textos = [w['text'] for w in ws]
            # Saltar encabezado de columnas y el pie de página que se repiten
            if HDR & set(textos):
                continue
            if any('descarga' in t.lower() or 'Banking' in t for t in textos):
                continue
            first = ws[0]
            if DATE4.match(first['text']) and first['x0'] < 90:
                # Nueva fila de movimiento
                if cur: movs.append(cur)
                desc = ' '.join(w['text'] for w in ws
                                if 100 <= w['x0'] < 300 and not RE_NUM.match(w['text'])
                                and w['text'] not in ('+', '-', '$')).rstrip(' -+').strip()
                signo = next((w['text'] for w in ws if w['text'] in ('+', '-')), '')
                nums = [(num_ar(w['text']), w['x1']) for w in ws if RE_NUM.match(w['text'])]
                der = [n for n in nums if n[1] > 460]          # saldo = número más a la derecha
                saldo = der[-1][0] if der else None
                otros = [n for n in nums if n[1] <= 460]
                imp = otros[0][0] if otros else 0.0
                deb = cre = 0.0
                if signo == '+':   cre = abs(imp)
                elif signo == '-': deb = abs(imp)
                else:                                           # respaldo por posición
                    if otros and otros[0][1] < 360: deb = abs(imp)
                    else:                           cre = abs(imp)
                cur = dict(fecha=first['text'], desc=desc, deb=deb, cre=cre, saldo=saldo, det=[])
            else:
                # Línea de detalle (nombre, CUIT, CBU, VARIOS, etc.) del movimiento en curso
                if cur is None: continue
                det = [w['text'] for w in ws if 100 <= w['x0'] < 300]
                if det: cur['det'].append(' '.join(det))
    if cur: movs.append(cur)

    out = []
    for m in movs:
        det = m['det']; nombre = det[0] if det else ''
        cu = next((d.strip() for d in det if _CUIT11.match(d.strip())), '')
        out.append(Movimiento(fecha_dmy4(m['fecha']), m['desc'], '', cu, nombre,
                              m['deb'], m['cre'], m['saldo']))
    periodo = ''
    if out:
        fs = [m.fecha for m in out if m.fecha]
        if fs: periodo = f"{min(fs).strftime('%d/%m/%Y')} a {max(fs).strftime('%d/%m/%Y')}"
    saldo_fin = out[-1].saldo if out and out[-1].saldo is not None else 0.0
    saldo_ini = round(out[0].saldo - (out[0].credito - out[0].debito), 2) if out and out[0].saldo is not None else 0.0
    return Resultado('Banco Galicia', '', '', cuenta, periodo, saldo_ini, saldo_fin, out)

# ---------------------------------------------------------------------------
# BBVA — 1 línea por movimiento, VARIAS CUENTAS por PDF. Débitos con signo.
# Columnas por borde derecho: débito x1<460 | crédito x1<540 | saldo resto.
# Cada cuenta: SALDO ANTERIOR ... movimientos ... SALDO AL / TOTAL MOVIMIENTOS.
# Se descartan los encabezados-resumen (no tienen SALDO ANTERIOR).
# La fecha viene dd/mm (sin año): se completa con el año del período.
# ---------------------------------------------------------------------------
def leer_bbva(pdf) -> List[Resultado]:
    DATE = re.compile(r'^\d{2}/\d{2}$')
    def zona(x1): return 'debito' if x1 < 460 else ('credito' if x1 < 540 else 'saldo')
    txt_all = '\n'.join((p.extract_text() or '') for p in pdf.pages)
    t0 = pdf.pages[0].extract_text() or ''
    mt = re.search(r'([A-ZÑ][A-ZÑ0-9&\.\- ]+?)\s*\((\d{2}-\d{8}-\d)\)', t0)
    titular = mt.group(1).strip() if mt else ''
    cuit_tit = mt.group(2) if mt else ''
    ma = re.search(r'(20\d\d)', t0); anio = ma.group(1) if ma else '2026'
    mp = re.search(r'MES:\s*([A-ZÑ]+ 20\d\d)', txt_all.upper()); periodo = mp.group(1).title() if mp else ''

    accts=[]; cur=None; inside=False
    for page in pdf.pages:
        for toks in _cluster(page.extract_words()):
            txt=' '.join(w['text'] for w in toks)
            m=re.search(r'(CC (?:\$|U\$S) [\d\-/]+)\s*\(Cta\.Cte', txt)
            if m:
                cur=dict(cuenta=m.group(1), saldo_ini=None, saldo_fin=None, movs=[])
                accts.append(cur); inside=False; continue
            if cur is None: continue
            up=txt.upper()
            if up.startswith('SALDO ANTERIOR'):
                a=[num_ar(w['text']) for w in toks if RE_NUM.match(w['text'])]
                if a: cur['saldo_ini']=a[-1]
                inside=True; continue
            if up.startswith('SALDO AL'):
                a=[num_ar(w['text']) for w in toks if RE_NUM.match(w['text'])]
                if a: cur['saldo_fin']=a[-1]
                inside=False; continue
            if up.startswith('TOTAL MOVIMIENTOS'):
                inside=False; continue
            first=toks[0]
            if inside and DATE.match(first['text']) and first['x0'] < 90 and first['text'] != '00/00':
                mv=dict(fecha=first['text'], origen='', concepto=[], debito=0.0, credito=0.0, saldo=None)
                for w in toks[1:]:
                    x0,x1,tx = w['x0'], w['x1'], w['text']
                    if RE_NUM.match(tx) and x1 > 380:
                        v=num_ar(tx); z=zona(x1)
                        if z=='debito': mv['debito']=abs(v)
                        elif z=='credito': mv['credito']=abs(v)
                        else: mv['saldo']=v
                    elif 90 <= x0 < 133: mv['origen'] += tx+' '
                    elif 133 <= x0 < 385: mv['concepto'].append(tx)
                mv['concepto']=' '.join(mv['concepto']).strip(); mv['origen']=mv['origen'].strip()
                cur['movs'].append(mv)

    res=[]
    for a in accts:
        if a['saldo_ini'] is None: continue     # encabezado-resumen duplicado
        movimientos=[Movimiento(_fecha_ddmm(mv['fecha'], anio), mv['concepto'], mv['origen'],
                                '', '', mv['debito'], mv['credito'], mv['saldo']) for mv in a['movs']]
        sfin = a['saldo_fin'] if a['saldo_fin'] is not None else (movimientos[-1].saldo if movimientos else a['saldo_ini'])
        res.append(Resultado('Banco BBVA', titular, cuit_tit, a['cuenta'], periodo,
                             a['saldo_ini'], sfin, movimientos))
    return res

# ---------------------------------------------------------------------------
# MACRO — monospace ancho fijo, VARIAS CUENTAS por PDF, cuenta cortada entre
# páginas (el encabezado de cuenta se repite: se mantiene el estado si es la
# misma). NO publica totales de movimientos -> control por continuidad de saldo.
# Columnas por borde derecho: débito x1<410 | crédito x1<490 | saldo resto.
# ---------------------------------------------------------------------------
def leer_macro(pdf) -> List[Resultado]:
    DATE = re.compile(r'^\d{2}/\d{2}/\d{2}$')
    ACC  = re.compile(r'CUENTA CORRIENTE.*NRO\.?:\s*([\d\-]+)')
    def zona(x1): return 'debito' if x1 < 410 else ('credito' if x1 < 490 else 'saldo')
    t0 = pdf.pages[0].extract_text() or ''
    titular=''
    lineas=[l.strip() for l in t0.split('\n')]
    for i,l in enumerate(lineas):
        if 'Sr(es)' in l:
            for j in range(i+1, min(i+4, len(lineas))):
                if lineas[j]: titular=lineas[j]; break
            break
    mc=re.search(r'C\.?U\.?I\.?T\.?\s*:?\s*(\d{11})', t0); cuit_tit=mc.group(1) if mc else ''
    mp=re.search(r'Periodo del Extracto:\s*([\d/]+ al [\d/]+)', t0); periodo=mp.group(1) if mp else ''

    accts=[]; cur=None; inside=False; cur_num=None
    for page in pdf.pages:
        for toks in _cluster(page.extract_words()):
            txt=' '.join(w['text'] for w in toks)
            m=ACC.search(txt)
            if m:
                num=m.group(1)
                if num != cur_num:
                    cur=dict(cuenta=num, saldo_ini=None, saldo_fin=None, movs=[])
                    accts.append(cur); cur_num=num; inside=False
                continue
            if cur is None: continue
            up=txt.upper()
            if 'SALDO ULTIMO EXTRACTO' in up:
                a=[num_ar(w['text']) for w in toks if RE_NUM.match(w['text'])]
                if a: cur['saldo_ini']=a[-1]
                inside=True; continue
            if 'SALDO FINAL AL DIA' in up:
                a=[num_ar(w['text']) for w in toks if RE_NUM.match(w['text'])]
                if a: cur['saldo_fin']=a[-1]
                inside=False; continue
            first=toks[0]
            if inside and DATE.match(first['text']) and first['x0'] < 80:
                mv=dict(fecha=first['text'], desc=[], ref='', debito=0.0, credito=0.0, saldo=None)
                for w in toks[1:]:
                    x0,x1,tx = w['x0'], w['x1'], w['text']
                    if RE_NUM.match(tx) and x1 > 320:
                        v=num_ar(tx); z=zona(x1)
                        if z=='debito': mv['debito']=abs(v)
                        elif z=='credito': mv['credito']=abs(v)
                        else: mv['saldo']=v
                    elif 67 <= x0 < 260: mv['desc'].append(tx)
                    elif 260 <= x0 < 339: mv['ref'] += tx+' '
                mv['desc']=' '.join(mv['desc']).strip(); mv['ref']=mv['ref'].strip()
                cur['movs'].append(mv)

    res=[]
    for a in accts:
        if not a['movs']: continue              # cuentas sin movimientos: se omiten
        movimientos=[Movimiento(fecha_dmy2(mv['fecha']), mv['desc'], mv['ref'],
                                '', '', mv['debito'], mv['credito'], mv['saldo']) for mv in a['movs']]
        sfin = a['saldo_fin'] if a['saldo_fin'] is not None else (movimientos[-1].saldo if movimientos else 0.0)
        res.append(Resultado('Banco Macro', titular, cuit_tit, a['cuenta'], periodo,
                             a['saldo_ini'] or 0.0, sfin, movimientos))
    return res

# ===========================================================================
# IMPUTACIÓN DE PROVEEDORES  (CUIT -> importe -> nombre)
# ---------------------------------------------------------------------------
# Usa el archivo de proveedores del cliente (listado + cuenta corriente).
#   1) Por CUIT: el extracto trae el CUIT de la transferencia -> proveedor.
#   2) Por importe exacto: el débito coincide con una factura de compra.
#   3) Si varios proveedores comparten importe: se desempata por nombre en el texto.
# Si no se identifica, queda "a imputar" para revisión humana (NUNCA se adivina).
# ===========================================================================
def cargar_proveedores(path: str) -> dict:
    """Lee el Excel de proveedores del cliente. Devuelve índices para imputar:
       por_cuit {cuit11 -> nombre}, por_importe {importe -> {nombres}}."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    por_cuit = {}; nombres = set(); por_importe = {}
    for name in wb.sheetnames:
        ws = wb[name]
        H = {str(ws.cell(1, c).value or '').strip().lower(): c
             for c in range(1, ws.max_column + 1)}
        # Hoja directorio: Nombre + Número de Identificación (CUIT)
        col_id = H.get('número de identificación') or H.get('numero de identificacion')
        if 'nombre' in H and col_id:
            for r in range(2, ws.max_row + 1):
                nom = ws.cell(r, H['nombre']).value; cid = ws.cell(r, col_id).value
                if nom and cid:
                    cu = re.sub(r'\D', '', str(cid))
                    if len(cu) == 11: por_cuit[cu] = str(nom).strip()
                    nombres.add(str(nom).strip().upper())
        # Hoja cuenta corriente: importes de Facturas de Compra por proveedor
        if 'proveedor' in H and 'haber' in H and 'documento' in H:
            for r in range(2, ws.max_row + 1):
                doc = str(ws.cell(r, H['documento']).value or '')
                prov = ws.cell(r, H['proveedor']).value
                hab = ws.cell(r, H['haber']).value
                if prov and hab and 'factura' in doc.lower():
                    por_importe.setdefault(round(float(hab), 2), set()).add(str(prov).strip().upper())
    return dict(por_cuit=por_cuit, por_importe=por_importe, nombres=nombres)

def imputar_proveedores(res: Resultado, prov: dict, cuenta: str = 'Proveedores') -> Resultado:
    """Identifica el proveedor de cada débito marcado 'PROVEEDOR (a imputar)'."""
    for m in res.movimientos:
        if m.cuenta_sugerida: continue
        if not (m.categoria or '').startswith('PROVEEDOR'): continue
        matched = None
        cu = re.sub(r'\D', '', m.cuit or '')
        if len(cu) == 11 and cu in prov['por_cuit']:
            matched = prov['por_cuit'][cu]                      # 1) por CUIT
        else:
            cand = prov['por_importe'].get(round(m.debito, 2))
            if cand and len(cand) == 1:
                matched = next(iter(cand))                      # 2) por importe exacto
            elif cand and len(cand) > 1:                        # 3) desempate por nombre
                txt = _texto_mov(m).upper()
                hits = [n for n in cand if n in txt]
                if len(hits) == 1: matched = hits[0]
        # El contador manda TODOS los pagos a proveedores a la cuenta 'Proveedores'.
        # El CUIT solo sirve para identificar cuál proveedor (detalle), no la cuenta.
        m.categoria = 'PROVEEDOR' if matched else 'PROVEEDOR (sin identificar)'
        m.cuenta_sugerida = cuenta
        if matched and not m.nombre: m.nombre = matched
    return res

# ===========================================================================
# IMPUTACIÓN DE VEP / IMPUESTOS AFIP  (por importe -> concepto -> cuenta)
# Los pagos de AFIP salen del extracto como un importe (sin decir qué impuesto).
# Se cruza contra el listado de VEP del cliente (hoja 'VEP' del liquidador:
# Importe + Descripcion, ej. 'IVA DJ01/26', 'SIJPDJ02/26', 'IIBBBA02/26').
# Si el mismo importe está en varios VEP, se desempata por fecha más cercana.
# ===========================================================================
def cargar_vep(path: str, hoja: str = 'VEP') -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if hoja not in wb.sheetnames:
        return dict(por_importe={})
    ws = wb[hoja]
    H = {str(ws.cell(1, c).value or '').strip().lower(): c for c in range(1, ws.max_column + 1)}
    cI = H.get('importe'); cD = H.get('descripcion') or H.get('descripción'); cF = H.get('fecha de pago')
    por_importe = {}
    if not cI:
        return dict(por_importe={})
    for r in range(2, ws.max_row + 1):
        imp = ws.cell(r, cI).value
        if imp is None: continue
        try: imp = round(float(imp), 2)
        except (TypeError, ValueError): continue
        desc = str(ws.cell(r, cD).value or '').strip() if cD else ''
        fec = ws.cell(r, cF).value if cF else None
        por_importe.setdefault(imp, []).append((desc, fec))
    return dict(por_importe=por_importe)

def _cuenta_impuesto(desc: str):
    d = (desc or '').upper()
    if 'IVA' in d:                                                 return ('IMPUESTO IVA', 'IVA a pagar')
    if 'SIJP' in d or 'SUSS' in d or 'SEG SOCIAL' in d:            return ('CARGAS SOCIALES', 'CARGAS SOCIALES A PAGAR')
    if any(k in d for k in ('IIBB', 'IBBA', 'BRUTOS', 'SIRCREB')): return ('INGRESOS BRUTOS', 'Ingresos brutos a pagar')
    if 'GANANCIA' in d:                                            return ('GANANCIAS', 'Impuesto a las Ganancias')
    if 'INTERES' in d or 'RESARC' in d:                            return ('INTERESES AFIP', 'Impuestos y Tasas')
    return ('IMPUESTO AFIP', 'Impuestos y Tasas')

def imputar_vep(res: Resultado, vep: dict, tol_dias: int = 45, cuenta_afip: str = None) -> Resultado:
    """Identifica el impuesto de cada pago de AFIP marcado 'IMPUESTOS AFIP/VEP'.
       Si se pasa cuenta_afip (ej. 'AFIP-RENTAS'), TODOS los pagos de AFIP van a esa
       cuenta puente (como hace el contador en la conciliación), y el impuesto puntual
       (SIJPDJ05/26, etc.) queda como detalle en 'nombre'. Si no, se separa por impuesto."""
    import datetime
    def _dif(item, fmov):
        f = item[1]
        if isinstance(f, datetime.datetime): f = f.date()
        if isinstance(f, datetime.date) and fmov: return abs((f - fmov).days)
        return 10 ** 6
    for m in res.movimientos:
        if m.cuenta_sugerida: continue
        if not (m.categoria or '').startswith('IMPUESTOS AFIP'): continue
        cand = vep['por_importe'].get(round(m.debito, 2))
        if not cand: continue
        if len(cand) == 1:
            elegido = cand[0]
        else:
            mejor = min(cand, key=lambda it: _dif(it, m.fecha))
            elegido = mejor if _dif(mejor, m.fecha) <= tol_dias else None
        if elegido:
            if cuenta_afip:
                m.categoria = 'IMPUESTO AFIP'; m.cuenta_sugerida = cuenta_afip
            else:
                m.categoria, m.cuenta_sugerida = _cuenta_impuesto(elegido[0])
            m.nombre = elegido[0]        # concepto VEP puntual (ej. SIJPDJ05/26)
    return res

# Firmas por texto EN MAYÚSCULAS. Ojo: el extracto de Galicia puede contener la
# palabra "BBVA" (en un detalle de transferencia), y en el PDF de BBVA el nombre
# y CUIT del banco vienen como imagen. Por eso se detecta por texto estructural
# propio de cada banco, no por el nombre del banco.
_FIRMAS = [
    ('Banco Comafi', leer_comafi,     lambda t: 'COMAFI' in t or '30-60473101-8' in t),
    ('Supervielle',  leer_supervielle,lambda t: 'SUPERVIELLE' in t or '33-50000517-9' in t or 'IAUREG010000' in t or 'AJBXP' in t),
    ('Banco Ciudad', leer_ciudad,     lambda t: '30-99903208-3' in t or 'BANCOCIUDAD' in t or 'BANCO CIUDAD' in t),
    ('Banco Macro',  leer_macro,      lambda t: 'MACRO' in t or '30-50001008-4' in t),
    ('Banco BBVA',   leer_bbva,       lambda t: 'CTA.CTE.BANCARIA' in t or 'CUENTA PYME' in t or 'SALDO ANTERIOR' in t),
    ('Banco Galicia',leer_galicia,    lambda t: 'RESUMEN DE CUENTA CORRIENTE' in t),
    ('Banco Galicia',leer_galicia_ob, lambda t: 'MOVIMIENTOS DE CC' in t and 'OFFICE BANKING' in t),
]

def detectar_banco(pdf):
    texto = "\n".join((pdf.pages[i].extract_text() or '') for i in range(min(2, len(pdf.pages)))).upper()
    for nombre, lector, test in _FIRMAS:
        if test(texto):
            return nombre, lector
    raise ValueError("No pude identificar el banco de este PDF. "
                     "Agregá su firma y su lector en el módulo.")

def leer_extracto(path: str) -> List[Resultado]:
    """Devuelve SIEMPRE una lista de Resultado (una por cuenta).
       Bancos de una sola cuenta devuelven una lista de un elemento."""
    pdf = pdfplumber.open(path)
    _, lector = detectar_banco(pdf)
    out = lector(pdf)
    return out if isinstance(out, list) else [out]

# ===========================================================================
# Control de integridad + exportación a Excel
# ===========================================================================
def verificar_control(res: Resultado):
    tot_d = sum(m.debito for m in res.movimientos)
    tot_c = sum(m.credito for m in res.movimientos)
    calc  = res.saldo_ini + tot_c - tot_d
    dif   = round(calc - res.saldo_fin, 2)
    acum, parciales = res.saldo_ini, 0
    for m in res.movimientos:
        acum += m.credito - m.debito
        if m.saldo is not None and abs(acum - m.saldo) > 0.01:
            parciales += 1
    return dict(total_debitos=tot_d, total_creditos=tot_c, saldo_calculado=calc,
                diferencia=dif, ok=(dif == 0 and parciales == 0),
                saldos_parciales_mal=parciales)

# ===========================================================================
# CLASIFICACIÓN CONTABLE (primer ladrillo: SUELDOS)
# ---------------------------------------------------------------------------
# Reglas de TEXTO (sin IA, auditables). Cada banco escribe distinto el pago de
# haberes, pero todos usan las palabras HABERES o SUELDO. Se detecta por esa
# palabra ancla (no por la frase exacta) para que aguante cambios de redacción
# del banco -> ej.: BBVA pasó de "DB/CR POR PAGO DE SUELDOS" a "DEBITO POR PAGO
# DE HABERES OL". Relevamiento por banco:
#   Galicia : "SERVICIO ACREDITAMIENTO DE HABERES" / "ACRED.HABERES"
#   Comafi  : "Transf inmed sueldos e-Banking"
#   BBVA    : "DEBITO POR PAGO DE HABERES OL" / "DB/CR POR PAGO DE SUELDOS"
#             (además origen "D 569" como señal de respaldo)
# La cuenta contable exacta depende del plan de cuentas del cliente: se pasa por
# parámetro. Por defecto, etiqueta genérica a confirmar contra el plan.
# ===========================================================================
CUENTA_SUELDOS_DEFAULT = 'Sueldos a pagar'

# Diccionario del contador (criterios de imputación aprendidos de un cliente real).
# Es la fuente de verdad por defecto; se puede pasar otro 'mapeo' para casos especiales.
MAPEO_CONTADOR = {
    'ACREDITACION MASTERCARD': 'Deudores por Venta',
    'ACREDITACION VISA': 'Deudores por Venta',
    'ANULACION DEBITOS': 'Proveedores',
    'ANULACION I LEY 25.413': 'Impuesto al Débito Ley 25.413',
    'ANULACION IMP LEY 25413 DEB 0,6%': 'Impuesto al Débito Ley 25.413',
    'ANULACION VISA': 'Proveedores',
    'CASHBACK BIENVENIDA PYME GALICIA': 'Impuestos y Tasas',
    'CHEQUE 48 HORAS': 'Proveedores',
    'CHEQUE DE CAJA': 'Caja',
    'CHEQUE PAGADOR POR CAJA': 'Proveedores',
    'CHEQUE RECHAZADO': 'Caja',
    'COM CHEQUE RECHAZADO': 'Gastos bancarios',
    'COM CONS CHEQ S/ SALDO': 'Gastos bancarios',
    'COM CONS CHEQ S\\ SALDO': 'Gastos bancarios',
    'COM MOVIMIENTOS OTRA SUCURSAL': 'Gastos bancarios',
    'COM VALOR AL COBRO': 'Gastos bancarios',
    'COM.GESTION COBRO CH': 'Gastos bancarios',
    'COM.IMP/SERBEPE': 'Gastos bancarios',
    'COM.POR EXTRACCIONES': 'Gastos bancarios',
    'COM.TRANSF.INTERNET': 'Gastos bancarios',
    'COMISION EXCESO MOVS EXTRACCIONES': 'Gastos bancarios',
    'COMISION MANT,CTA': 'Gastos bancarios',
    'COMISION MANT.CTA': 'Gastos bancarios',
    'COMISION MOV CLEARING': 'Gastos bancarios',
    'COMISION POR CHEQUERA': 'Gastos bancarios',
    'COMISION POR SERVICIO DE CUENTA': 'Gastos bancarios',
    'COMISION SERVICIO DE CUENTA': 'Gastos bancarios',
    'CR COMER': 'Proveedores',
    'CRED.ACOMERCIOS TARJ.CRED VISA': 'Proveedores',
    'CREDITO TRANSFERENCIA': 'Deudores por Venta',
    'CREDITO TRANSFERENCIA COELSA': 'Deudores por Venta',
    'DB.COMERCIO': 'Proveedores',
    'DEB.COMERCIO TC': 'Proveedores',
    'DEBITO POR COMPRA VENTA DOLARES - USD': 'Moneda Extranjera',
    'DEBITO TRANSF. ONLINE BANKING EMP - A CTA 0217-007003612853 ARS': 'Proveedores',
    'DEP CH 24 HS AUTOSERV': 'Caja',
    'DEP. CANJE INTERNO': 'Caja',
    'DEP.CHEQUES AUTO': 'Caja',
    'DEP.CHEQUES AUTOSERV': 'Caja',
    'DEPOSITO CHEQUES': 'Caja',
    'DEPOSITO EFVO CAJ AUT': 'Caja',
    'DEPOSITO INTERSUCURSAL (CAJA)': 'Caja',
    'DEV.IMP.DEB.LEY 25413-ALIC.GENERAL': 'Impuesto al Débito Ley 25.413',
    'DGI/AFIP': 'AFIP-RENTAS',
    'FCI SUSCRIPCION': 'FCI',
    'HONORARIOS DE PROFESIONALES': 'Deudores por Venta',
    'I,V,A,': 'IVA Crédito Fiscal',
    'I.V.A': 'IVA Crédito Fiscal',
    'I.V.A. REDUCIDO 50%': 'IVA Crédito Fiscal',
    'IB. MULTIL.CABA': 'Retenciones Ingresos Brutos CABA',
    'IBCABA PERCEP': 'Percepción Ingresos Brutos Sufrida',
    'IMP LEY 25413 CRED 0,6%': 'Impuesto al Crédito Ley 25.413',
    'IMP SELLOS CABA AD. EN CTA.CTE': 'Gastos bancarios',
    'IMP. CRE. LEY 25413': 'Impuesto al Crédito Ley 25.413',
    'IMP. DEB. LEY 25413 GRAL': 'Impuesto al Débito Ley 25.413',
    'IMP. ING. BRUTOS': 'Percepción Ingresos Brutos Sufrida',
    'IMPUESTO DE SELLOS': 'Impuestos y Tasas',
    'ING. BRUTOS S/ CRED': 'Percepción Ingresos Brutos Sufrida',
    'INGRESOS BRUTOS': 'Impuestos y Tasas',
    'INTERES SALDO DEUDOR': 'Gastos bancarios',
    'INTERESES POR DESCUBIERTO': 'Gastos bancarios',
    'INTERESES SOBRE SALDOS DEUDORES': 'Gastos bancarios',
    'IVA': 'IVA Crédito Fiscal',
    'IVA - REDUCIDO 10,5%': 'IVA Crédito Fiscal',
    'IVA PERCEPCION': 'Percepción de IVA Sufrida',
    'IVA TASA GENERAL': 'IVA Crédito Fiscal',
    'MANTENIMIENTO': 'Gastos bancarios',
    'MUNI CIUDAD BA': 'Impuestos y Tasas',
    'NAVE - VENTA CON TARJETA': 'Deudores por Venta',
    'NAVE PAGO CON TRANSFERENCIA': 'Deudores por Venta',
    'PAGO A COMERCIOS': 'Deudores por Venta',
    'PAGO A COMERCIOS VISA': 'Deudores por Venta',
    'PAGO A PROVEEDORES - 180728078MERCADOLIBRE SRL': 'Deudores por Venta',
    'PAGO CCI 24HS GRAVADA INTERBANKING - A CBU 0070105720000005696688': 'Proveedores',
    'PAGO CHEQ.DE MOSTRADOR(CAJA)': 'Proveedores',
    'PAGO DE CH.INTERSUCURSAL(CAJA)': 'Proveedores',
    'PAGO PROVEEDORES DATANET': 'Proveedores',
    'PCT PAGO CON TRANSF': 'Deudores por Venta',
    'PERC.INGRESOS BRUTOS': 'Percepción Ingresos Brutos Sufrida',
    'PERCEP. IVA': 'Percepción de IVA Sufrida',
    'RECAUDACIONES ELECTRONICAS TARJ': 'Proveedores',
    'REG.REC.SIRCREB': 'Sircreb',
    'RESUMEN DE CUENTA': 'Gastos bancarios',
    'RETIRO EN EFVO POR CAJA SUC 0395': 'Caja',
    'SELLADO': 'Gastos bancarios',
    'SERV. AGUA/ GAS': 'Agua',
    'SERVIC ELECTRIC': 'Energía Eléctrica',
    'SERVIC TELEFONI': 'Telefonia',
    'SERVIC/IMPUESTO': 'AFIP-RENTAS',
    'SERVICIO ACREDITAMIENTO DE HABERES': 'SUELDOS A PAGAR',
    'SERVICIO PAGO A PROVEEDORES': 'Deudores por Venta',
    'SERVICIO TERMINAL PAYWAY': 'Proveedores',
    'SNP PAGO A PROVEEDORES': 'Deudores por Venta',
    'SUELDOS': 'Deudores por Venta',
    'SUELDOS TRANSFERIDOS': 'SUELDOS A PAGAR',
    'SUSCRIPCION FIMA': 'FCI',
    'TARJ,CRED,(AJUSTE COM)': 'Gastos bancarios',
    'TARJ.CRED.(AJUSTE COM)': 'Gastos bancarios',
    'TRANSF. AFIP': 'EATON DIEGO MARTIN - Cuenta Particular',
    'TRANSF. CTAS PROPIAS': 'Banco Galicia en $',
    'TRANSF.DATANET': 'Deudores por Venta',
    'TRANSF.ELECTRON': 'Deudores por Venta',
    'TRANSF.FONDOS': 'Proveedores',
    'TRANSFER. CASH MISMA TITULARIDAD': 'Banco Galicia en $',
    'TRANSFERENCIA': 'Proveedores',
    'TRANSFERENCIA DE CUENTA PROPIA': 'Banco Galicia en $',
    'TRANSFERENCIA DE TERCEROS': 'Deudores por Venta',
    'TRANSFERENCIAS': 'Deudores por Venta',
    'TRANSFERENCIAS CASH PROVEEDORES': 'Deudores por Venta',
    'TRASPASO DE SALDO G+': 'Banco',
    'TRF ORDEN JUDIC': 'Proveedores',
}


# Reglas de clasificación por TEXTO del concepto. El orden es la prioridad: gana
# la primera que coincide. La 'cuenta' es una ETIQUETA por defecto y hay que
# confirmarla contra el plan de cuentas de cada cliente. cuenta '' = todavía no
# se puede resolver solo: requiere el paso siguiente (VEP por importe, o
# proveedor por CUIT->importe->nombre) -> queda marcado "a imputar".
# (categoria, cuenta_sugerida, patrón, aplica)  aplica: 'debito' | 'credito' | 'ambos'
REGLAS_CONCEPTO = [
    ('SUELDOS',              CUENTA_SUELDOS_DEFAULT,           r'HABERES|SUELDO',                                'debito'),
    ('IMP. LEY 25413 CRED',  'Impuesto al Crédito Ley 25.413', r'(?:IMP\.?\s*CRE|SOBRE\s*CRED).*25\.?413|25\.?413.*CRE|IMPUESTO A LOS CREDITOS', 'ambos'),
    ('IMP. LEY 25413 DEB',   'Impuesto al Débito Ley 25.413',  r'25\.?413|DEBITOS Y CREDITOS|IMPUESTO A LOS DEBITOS|TRANSFINAN|TRANSACCIONES FINAN', 'ambos'),
    ('PERCEPCION IVA',       'Percepción de IVA Sufrida',      r'PERCEP\w*\.?\s*IVA|PERC\.?\s*IVA',               'debito'),
    ('SIRCREB',              'Sircreb',                        r'SIRCREB',                                       'debito'),
    ('PERCEPCION IIBB',      'Percepción Ingresos Brutos Sufrida', r'ING\.?\s*BRUT|INGRESOS BRUTOS|IIBB|PERC\w*\.?\s*(?:ING|CABA)|CABA\s*ING', 'debito'),
    ('IMPUESTO SELLOS',      'Impuestos y Tasas',              r'SELLOS|SELLADO',                                'debito'),
    ('TRANSF. CTAS PROPIAS', 'Transferencias entre cuentas',   r'CCP\d|CAP\d|CTAS?\s*PROP|MISMA TITULARIDAD|CUENTAS? PROPIAS?|CUENTA PROPIA', 'ambos'),
    ('COMISIONES',           'Gastos bancarios',               r'COMISION|COMI\b|COM\s*MANT|MANTENIMIENTO|SERVICIO DE CUENTA|COMIS\b', 'debito'),
    ('IVA',                  'IVA Crédito Fiscal',             r'\bIVA\b',                                       'debito'),
    ('INTERESES',            'Gastos bancarios',               r'INTERES',                                       'debito'),
    ('IMPUESTOS AFIP/VEP',   '',                               r'IMP\.?\s*AFIP|PAGOS?\s*AFIP|SERVICIOS IMP',     'debito'),
    ('PROVEEDOR',            'Proveedores',                    r'PROVEEDOR|TRF\s*INMED|TRANSFER|TRANSF',         'debito'),
    ('COBRANZA',             None,                             r'.',                                             'credito'),
]
_REGLAS = [(cat, cta, re.compile(pat, re.IGNORECASE), ap) for cat, cta, pat, ap in REGLAS_CONCEPTO]

import unicodedata as _ud
def _sin_acentos(s: str) -> str:
    return ''.join(c for c in _ud.normalize('NFD', str(s)) if _ud.category(c) != 'Mn')

def _txt_regla(m) -> str:
    """Texto del movimiento normalizado para las reglas: sin acentos ni puntos, MAYÚSCULAS.
       Así 'I.V.A.' -> 'IVA' y 'Débitos y Créditos' -> 'DEBITOS Y CREDITOS'."""
    return _sin_acentos(f"{m.concepto} {m.nombre} {m.referencia}").upper().replace('.', '')

def _texto_mov(m: Movimiento) -> str:
    # En Galicia el nombre/detalle (p.ej. 'HABERES') va en 'nombre'; se incluye.
    return f"{m.concepto} {m.nombre} {m.referencia}"

def _norm_concepto(s: str) -> str:
    """Normaliza un concepto para cruzar extracto vs tabla del cliente."""
    s = re.sub(r'\s+', ' ', str(s or '').upper().strip())
    return s.rstrip(' .')

def cargar_mapeo_conceptos(path: str, hoja: str = 'Concepto gasto y asignacion cta') -> dict:
    """Lee la tabla concepto->cuenta del liquidador del cliente (fuente de verdad).
       Solo usa conceptos con cuenta ÚNICA; los ambiguos (varias cuentas posibles)
       NO se auto-asignan (requieren CUIT/importe/humano)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if hoja not in wb.sheetnames:
        return dict(mapeo={}, ambiguos=set())
    ws = wb[hoja]
    H = {str(ws.cell(1, c).value or '').strip().upper(): c for c in range(1, ws.max_column + 1)}
    cB = H.get('CONCEPTO EXTRACTO'); cC = H.get('CUENTA CONTABLE')
    if not cB or not cC:
        return dict(mapeo={}, ambiguos=set())
    tmp = {}
    for r in range(2, ws.max_row + 1):
        con = ws.cell(r, cB).value; cta = ws.cell(r, cC).value
        if con and cta:
            tmp.setdefault(_norm_concepto(con), set()).add(str(cta).strip())
    mapeo = {k: next(iter(v)) for k, v in tmp.items() if len(v) == 1}
    ambiguos = {k for k, v in tmp.items() if len(v) > 1}
    return dict(mapeo=mapeo, ambiguos=ambiguos)

def es_sueldo(mov: Movimiento) -> bool:
    """True si el movimiento es un pago de sueldos/haberes."""
    return bool(re.search(r'HABERES|SUELDO', _texto_mov(mov), re.IGNORECASE)) and mov.debito > 0

def clasificar(res: Resultado, cuenta_sueldos: str = CUENTA_SUELDOS_DEFAULT, mapeo: dict = None,
               cuenta_cobranzas: str = 'Deudores por Venta') -> Resultado:
    """Asigna cuenta contable a cada movimiento.
       1) Tabla del cliente (fuente de verdad): prueba el concepto, y si no, concepto+nombre
          (en Galicia el concepto viene partido, ej. 'TRANSFERENCIA DE CUENTA' + 'PROPIA').
       2) Si no, REGLAS_CONCEPTO (respaldo). Todo crédito no identificado -> cuenta_cobranzas
          ('Deudores por venta'), como hace el contador en el asiento de COBRANZAS."""
    tabla = (mapeo.get('mapeo', {}) if mapeo else MAPEO_CONTADOR)  # por defecto, diccionario del contador
    for m in res.movimientos:
        if m.categoria: continue
        if tabla:
            for k in (_norm_concepto(m.concepto), _norm_concepto(f"{m.concepto} {m.nombre}")):
                if k in tabla:
                    m.categoria = 'MAPEO'; m.cuenta_sugerida = tabla[k]
                    break
            if m.categoria: continue
        txt = _txt_regla(m); es_deb = m.debito > 0; es_cred = m.credito > 0
        for cat, cta, rx, ap in _REGLAS:
            if ap == 'debito' and not es_deb: continue
            if ap == 'credito' and not es_cred: continue
            if rx.search(txt):
                if cat == 'SUELDOS':
                    m.categoria = cat; m.cuenta_sugerida = cuenta_sueldos
                elif cat == 'COBRANZA':
                    m.categoria = 'COBRANZA'; m.cuenta_sugerida = cuenta_cobranzas
                else:
                    m.categoria = cat; m.cuenta_sugerida = cta
                break
    return res

def _sanitizar_hoja(nombre, usados):
    for ch in '\\/?*[]:':
        nombre = nombre.replace(ch, '-')
    nombre = (nombre or 'Cuenta')[:28] or 'Cuenta'
    base=nombre; i=2
    while nombre in usados:
        nombre=f"{base[:25]}_{i}"; i+=1
    usados.add(nombre)
    return nombre

def _escribir_hoja(ws, res):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    AR=Font(name='Arial',size=10); ARB=Font(name='Arial',size=10,bold=True)
    TIT=Font(name='Arial',size=12,bold=True,color='FFFFFF'); HF=Font(name='Arial',size=10,bold=True,color='FFFFFF')
    FILL=PatternFill('solid',fgColor='1F4E78'); GREEN=PatternFill('solid',fgColor='C6EFCE')
    YEL=PatternFill('solid',fgColor='FFF2CC')
    MONEY='#,##0.00;[Red]-#,##0.00'; BD=Border(bottom=Side(style='thin',color='D9D9D9')); CEN=Alignment(horizontal='center')
    BLUE=PatternFill('solid',fgColor='DDEBF7'); ORANGE=PatternFill('solid',fgColor='FCE4D6')
    ws['A1']=f"{res.banco.upper()} — {res.titular} (CUIT {res.cuit_titular}) · Cta {res.cuenta} · {res.periodo}"
    ws['A1'].font=TIT; ws['A1'].fill=FILL; ws.merge_cells('A1:K1')
    ws['A1'].alignment=Alignment(horizontal='left',vertical='center'); ws.row_dimensions[1].height=20
    cols=['Fecha','Concepto','Referencia','CUIT','Nombre','Cuenta sugerida',
          'Débitos','Créditos','Saldo (banco)','Saldo calculado','Control']
    for j,c in enumerate(cols,1):
        cc=ws.cell(3,j,c); cc.font=HF; cc.fill=FILL; cc.alignment=CEN
    r=4; first=r
    for m in res.movimientos:
        if m.fecha is not None: ws.cell(r,1,m.fecha).number_format='dd/mm/yyyy'
        ws.cell(r,2,m.concepto); ws.cell(r,3,m.referencia); ws.cell(r,4,m.cuit); ws.cell(r,5,m.nombre)
        etiqueta = m.cuenta_sugerida or m.categoria
        cs=ws.cell(r,6,etiqueta)
        if m.cuenta_sugerida: cs.fill=BLUE
        elif m.categoria: cs.fill=ORANGE
        ws.cell(r,7,m.debito or None).number_format=MONEY
        ws.cell(r,8,m.credito or None).number_format=MONEY
        if m.saldo is not None: ws.cell(r,9,m.saldo).number_format=MONEY
        ws.cell(r,10, f'=$N$2+H{r}-G{r}' if r==first else f'=J{r-1}+H{r}-G{r}').number_format=MONEY
        ws.cell(r,11, f'=IF(I{r}="","",I{r}-J{r})').number_format=MONEY
        for j in range(1,12): ws.cell(r,j).font=AR; ws.cell(r,j).border=BD
        r+=1
    last=r-1
    n_cuit=sum(1 for m in res.movimientos if m.cuit)
    n_cta =sum(1 for m in res.movimientos if m.cuenta_sugerida)
    n_imp =sum(1 for m in res.movimientos if m.categoria and not m.cuenta_sugerida)
    n_sin =sum(1 for m in res.movimientos if not m.categoria)
    filas=[('Saldo inicial',res.saldo_ini),('Total Débitos',f'=SUM(G{first}:G{last})'),
           ('Total Créditos',f'=SUM(H{first}:H{last})'),('Saldo final calculado','=N2+N4-N3'),
           ('Saldo final real (banco)',res.saldo_fin),('CONTROL (debe dar 0)','=N5-N6'),
           ('Cant. movimientos',f'=COUNT(A{first}:A{last})'),
           ('Con cuenta sugerida',n_cta),('A imputar (revisar)',n_imp),('Sin clasificar',n_sin)]
    for i,(lbl,val) in enumerate(filas):
        rr=2+i; ws.cell(rr,13,lbl).font=ARB
        c=ws.cell(rr,14,val); c.font=AR; c.number_format=MONEY if i<6 else '0'
    ws['N2'].fill=YEL; ws['N6'].fill=YEL; ws['N7'].fill=GREEN
    for j,w in enumerate([11,32,13,13,20,22,15,15,15,15,12],1):
        ws.column_dimensions[get_column_letter(j)].width=w
    ws.column_dimensions['M'].width=24; ws.column_dimensions['N'].width=16
    ws.freeze_panes='A4'

def exportar_excel(resultados, path):
    """Acepta un Resultado o una lista. Escribe UNA HOJA por cuenta."""
    import openpyxl
    if not isinstance(resultados, list): resultados=[resultados]
    wb=openpyxl.Workbook(); primera=True; usados=set()
    for res in resultados:
        clasificar(res)                      # rellena cuenta_sugerida (SUELDOS por ahora)
        ws = wb.active if primera else wb.create_sheet(); primera=False
        ws.title=_sanitizar_hoja(res.cuenta, usados)
        _escribir_hoja(ws, res)
    wb.save(path)

# ===========================================================================
# CLI
# ===========================================================================
if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python conversor_extractos.py archivo.pdf [salida.xlsx]"); sys.exit(1)
    entrada = sys.argv[1]
    resultados = leer_extracto(entrada)
    print(f"Banco   : {resultados[0].banco}")
    print(f"Titular : {resultados[0].titular}  (CUIT {resultados[0].cuit_titular})")
    print(f"Cuentas : {len(resultados)}")
    todo_ok=True
    for res in resultados:
        ctrl=verificar_control(res); ok=ctrl['ok']; todo_ok = todo_ok and ok
        print(f"  · Cta {res.cuenta:22} {len(res.movimientos):3} movs | "
              f"saldo {res.saldo_ini:,.2f} -> {res.saldo_fin:,.2f} | "
              f"control {ctrl['diferencia']:,.2f}  {'OK ✓' if ok else 'REVISAR ✗'}")
    import os
    salida = sys.argv[2] if len(sys.argv) > 2 else os.path.basename(entrada).rsplit('.',1)[0] + '_convertido.xlsx'
    exportar_excel(resultados, salida)
    print(f"Excel   : {salida}   ({'TODO OK' if todo_ok else 'HAY CUENTAS A REVISAR'})")
